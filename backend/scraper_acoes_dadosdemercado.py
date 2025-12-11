"""
Script para extrair lista completa de ações do Dados de Mercado
URL: https://www.dadosdemercado.com.br/acoes
"""

import requests
from bs4 import BeautifulSoup
import re
from typing import List, Dict, Set
import time
import os
from datetime import datetime
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[AVISO] openpyxl não está instalado. Instale com: pip install openpyxl")

def extrair_tickers_de_pagina_dadosdemercado(soup: BeautifulSoup) -> Set[str]:
    """Extrai tickers de uma página HTML do Dados de Mercado"""
    tickers = set()
    
    # Estratégia 1: Buscar links que contenham tickers
    links = soup.find_all('a', href=True)
    for link in links:
        href = link.get('href', '')
        # Padrão: pode ter /acoes/TICKER ou /acao/TICKER ou similar
        match = re.search(r'/(?:acoes|acao)/([A-Z]{4}[0-9]{1,2})', href, re.IGNORECASE)
        if match:
            ticker = match.group(1).upper()
            # Validar formato de ação (4 letras + número)
            if re.match(r'^[A-Z]{4}[0-9]{1,2}$', ticker):
                # Ignorar variações fracionadas (terminam com F)
                if not ticker.endswith('F'):
                    tickers.add(ticker)
    
    # Estratégia 2: Buscar em tabelas
    tabelas = soup.find_all('table')
    for tabela in tabelas:
        linhas = tabela.find_all('tr')
        for linha in linhas:
            celulas = linha.find_all(['td', 'th'])
            for celula in celulas:
                texto = celula.get_text(strip=True)
                # Procurar padrão de ticker no texto
                match = re.search(r'\b([A-Z]{4}[0-9]{1,2})\b', texto)
                if match:
                    ticker = match.group(1).upper()
                    if re.match(r'^[A-Z]{4}[0-9]{1,2}$', ticker):
                        # Ignorar variações fracionadas
                        if not ticker.endswith('F'):
                            tickers.add(ticker)
    
    # Estratégia 3: Buscar em divs, spans e outros elementos com texto
    elementos = soup.find_all(['div', 'span', 'p', 'td', 'th', 'a', 'li'])
    for elem in elementos:
        texto = elem.get_text(strip=True)
        # Procurar padrão de ticker (4 letras + 1-2 números)
        matches = re.findall(r'\b([A-Z]{4}[0-9]{1,2})\b', texto)
        for match in matches:
            ticker = match.upper()
            if re.match(r'^[A-Z]{4}[0-9]{1,2}$', ticker):
                # Ignorar variações fracionadas
                if not ticker.endswith('F'):
                    tickers.add(ticker)
    
    
    elementos_data = soup.find_all(attrs={'data-ticker': True})
    elementos_data.extend(soup.find_all(attrs={'data-symbol': True}))
    for elem in elementos_data:
        ticker = (elem.get('data-ticker') or elem.get('data-symbol', '')).upper()
        if re.match(r'^[A-Z]{4}[0-9]{1,2}$', ticker):
            if not ticker.endswith('F'):
                tickers.add(ticker)
    
    # Estratégia 5: Buscar em elementos com classes específicas de ticker
    elementos_ticker = soup.find_all(class_=lambda x: x and ('ticker' in str(x).lower() or 'symbol' in str(x).lower() or 'codigo' in str(x).lower()))
    for elem in elementos_ticker:
        texto = elem.get_text(strip=True)
        match = re.search(r'\b([A-Z]{4}[0-9]{1,2})\b', texto)
        if match:
            ticker = match.group(1).upper()
            if re.match(r'^[A-Z]{4}[0-9]{1,2}$', ticker):
                if not ticker.endswith('F'):
                    tickers.add(ticker)
    
    return tickers


def detectar_total_paginas_dadosdemercado(soup: BeautifulSoup) -> int:
    """Tenta detectar o número total de páginas no Dados de Mercado"""
    max_pagina = 1
    
    # Estratégia 1: Procurar por links de paginação
    links_paginacao = soup.find_all('a', href=re.compile(r'[?&]page=\d+|/page/\d+'))
    for link in links_paginacao:
        href = link.get('href', '')
        # Procurar page= ou /page/
        match = re.search(r'(?:[?&]page=|/page/)(\d+)', href)
        if match:
            try:
                num = int(match.group(1))
                if 1 <= num <= 100:  # Validar
                    max_pagina = max(max_pagina, num)
            except:
                pass
    
    # Estratégia 2: Procurar em botões ou spans de paginação
    elementos_paginacao = soup.find_all(['button', 'span', 'div', 'a'], 
                                       class_=lambda x: x and ('page' in str(x).lower() or 'pagination' in str(x).lower() or 'pager' in str(x).lower()))
    for elem in elementos_paginacao:
        texto = elem.get_text(strip=True)
        # Procurar números que parecem ser números de página
        match = re.search(r'\b(\d+)\b', texto)
        if match:
            try:
                num = int(match.group(1))
                if 1 <= num <= 100:  # Validar
                    max_pagina = max(max_pagina, num)
            except:
                pass
    
    # Estratégia 3: Procurar por "última página" ou "total de páginas"
    texto = soup.get_text()
    patterns = [
        r'(?:última|last|final)\s*(?:página|page)[:\s]*(\d+)',
        r'(?:total|de)\s*(?:páginas|pages)[:\s]*(\d+)',
        r'página\s*(\d+)\s*de\s*(\d+)',  # "página X de Y"
    ]
    for pattern in patterns:
        match = re.search(pattern, texto, re.IGNORECASE)
        if match:
            try:
                num = int(match.group(-1))  # Último grupo
                if 1 <= num <= 100:  # Validar
                    max_pagina = max(max_pagina, num)
            except:
                pass
    
    return max_pagina


def extrair_acoes_dadosdemercado() -> Dict[str, any]:
    """Extrai lista completa de ações do Dados de Mercado"""
    url_base = 'https://www.dadosdemercado.com.br/acoes'
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Referer': 'https://www.dadosdemercado.com.br/',
        'Cache-Control': 'no-cache',
        'Pragma': 'no-cache'
    }
    
    # Criar sessão com retry
    session = requests.Session()
    session.headers.update(headers)
    
    tickers = set()
    paginas_processadas = 0
    
    try:
        # Primeiro, buscar a página 1 para detectar o total de páginas
        # Pode ser /acoes ou /acoes?page=1 ou /acoes/page/1
        url_pagina1 = url_base
        print(f"[INFO] Buscando página 1: {url_pagina1}")
        
        # Tentar com retry
        max_tentativas = 3
        for tentativa in range(max_tentativas):
            try:
                response = session.get(url_pagina1, timeout=30)
                response.raise_for_status()
                break
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
                if tentativa < max_tentativas - 1:
                    print(f"  ⚠ Tentativa {tentativa + 1} falhou (timeout). Tentando novamente em 3 segundos...")
                    time.sleep(3)
                else:
                    raise
        
        soup = BeautifulSoup(response.content, 'html.parser')
        tickers_pagina = extrair_tickers_de_pagina_dadosdemercado(soup)
        tickers.update(tickers_pagina)
        paginas_processadas += 1
        print(f"  ✓ Página 1: {len(tickers_pagina)} tickers encontrados (total acumulado: {len(tickers)})")
        
        # Tentar detectar o número total de páginas
        max_pagina_detectada = detectar_total_paginas_dadosdemercado(soup)
        print(f"[INFO] Máxima página detectada na primeira página: {max_pagina_detectada}")
        
        # Limitar a um máximo razoável
        max_pagina_limite = min(max(max_pagina_detectada, 10), 50)  # Máximo de 50 páginas
        
        if max_pagina_detectada <= 1:
            print("[INFO] Não foi possível detectar o total de páginas. Tentando até 50 páginas...")
        else:
            print(f"[INFO] Usando limite de {max_pagina_limite} páginas")
        
        # Buscar todas as páginas
        # Tentar diferentes formatos de URL de paginação
        pagina_atual = 2
        paginas_sem_resultados = 0
        max_paginas_sem_resultados = 2  # Parar após 2 páginas consecutivas sem resultados
        
        while pagina_atual <= max_pagina_limite:
            # Tentar diferentes formatos de URL
            urls_tentativas = [
                f"{url_base}?page={pagina_atual}",
                f"{url_base}/page/{pagina_atual}",
                f"{url_base}/page/{pagina_atual}/",
            ]
            
            url_pagina = None
            response = None
            
            for url_tentativa in urls_tentativas:
                print(f"[INFO] Buscando página {pagina_atual}: {url_tentativa}")
                
                try:
                    # Tentar com retry
                    max_tentativas = 2
                    for tentativa in range(max_tentativas):
                        try:
                            response = session.get(url_tentativa, timeout=30)
                            if response.status_code == 200:
                                url_pagina = url_tentativa
                                break
                        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
                            if tentativa < max_tentativas - 1:
                                time.sleep(2)
                            else:
                                raise
                    
                    if url_pagina:
                        break
                        
                except requests.exceptions.HTTPError as e:
                    if e.response.status_code == 404:
                        # Tentar próximo formato de URL
                        continue
                    else:
                        raise
                except Exception as e:
                    # Tentar próximo formato de URL
                    continue
            
            if not url_pagina or not response:
                print(f"  ⚠ Não foi possível acessar página {pagina_atual} em nenhum formato. Parando.")
                break
            
            # Se retornar 404, não há mais páginas
            if response.status_code == 404:
                print(f"  ✗ Página {pagina_atual} não encontrada (404). Parando.")
                break
            
            try:
                response.raise_for_status()
                
                soup = BeautifulSoup(response.content, 'html.parser')
                tickers_pagina = extrair_tickers_de_pagina_dadosdemercado(soup)
                
                # Verificar se a página realmente tem conteúdo diferente
                if len(tickers_pagina) == 0:
                    paginas_sem_resultados += 1
                    print(f"  ⚠ Página {pagina_atual}: Nenhum ticker encontrado ({paginas_sem_resultados}/{max_paginas_sem_resultados})")
                    
                    if paginas_sem_resultados >= max_paginas_sem_resultados:
                        print(f"  ✗ {max_paginas_sem_resultados} páginas consecutivas sem resultados. Parando.")
                        break
                else:
                    # Verificar se são tickers novos ou já foram encontrados antes
                    tickers_novos = tickers_pagina - tickers
                    if len(tickers_novos) == 0 and len(tickers_pagina) > 0:
                        # Todos os tickers já foram encontrados, pode ser que não haja mais páginas
                        paginas_sem_resultados += 1
                        print(f"  ⚠ Página {pagina_atual}: {len(tickers_pagina)} tickers encontrados, mas todos já estavam na lista anterior")
                        if paginas_sem_resultados >= max_paginas_sem_resultados:
                            print(f"  ✗ {max_paginas_sem_resultados} páginas consecutivas sem novos tickers. Parando.")
                            break
                    else:
                        paginas_sem_resultados = 0  # Reset contador
                        tickers.update(tickers_pagina)
                        paginas_processadas += 1
                        print(f"  ✓ Página {pagina_atual}: {len(tickers_pagina)} tickers encontrados ({len(tickers_novos)} novos) (total acumulado: {len(tickers)})")
                
                time.sleep(1)  # Pausa entre requisições para não sobrecarregar o servidor
                
            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 404:
                    print(f"  ✗ Página {pagina_atual} não encontrada (404). Parando.")
                    break
                else:
                    print(f"  ⚠ Erro HTTP {e.response.status_code} na página {pagina_atual}: {e}")
                    paginas_sem_resultados += 1
                    if paginas_sem_resultados >= max_paginas_sem_resultados:
                        break
            except Exception as e:
                print(f"  ⚠ Erro ao processar página {pagina_atual}: {e}")
                paginas_sem_resultados += 1
                if paginas_sem_resultados >= max_paginas_sem_resultados:
                    break
            
            pagina_atual += 1
        
        # Converter para lista ordenada e adicionar .sa (minúsculo)
        tickers_list = sorted([f"{t}.sa" for t in tickers])
        
        print(f"\n[INFO] Resumo: {paginas_processadas} páginas processadas, {len(tickers_list)} tickers únicos encontrados")
        
        return {
            'sucesso': True,
            'tipo': 'acoes',
            'total': len(tickers_list),
            'tickers': tickers_list,
            'paginas_processadas': paginas_processadas,
            'url_base': url_base
        }
        
    except Exception as e:
        import traceback
        return {
            'sucesso': False,
            'erro': str(e),
            'traceback': traceback.format_exc(),
            'url_base': url_base
        }


def gerar_excel_comparacao(comparacao: Dict, tickers_novos: List[str], arquivo_saida: str = None) -> str:
    """Gera arquivo Excel com a comparação entre lista atual e nova"""
    if not OPENPYXL_AVAILABLE:
        print("[ERRO] openpyxl não está disponível. Não é possível gerar Excel.")
        return None
    
    if arquivo_saida is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo_saida = f"comparacao_acoes_{timestamp}.xlsx"
    
    # Criar workbook
    wb = Workbook()
    
    # Remover planilha padrão
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Aba 1: Ativos em Comum
    ws_iguais = wb.create_sheet("Ativos em Comum")
    ws_iguais.append(["Ticker"])
    ws_iguais.append(["Total de ativos em comum: " + str(len(comparacao['iguais']))])
    ws_iguais.append([])
    ws_iguais.append(["Ticker"])
    
    # Formatar cabeçalho
    for cell in ws_iguais[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Adicionar tickers em comum
    for ticker in sorted(comparacao['iguais']):
        ws_iguais.append([f"{ticker}.sa"])
    
    # Ajustar largura da coluna
    ws_iguais.column_dimensions['A'].width = 15
    
    # Aba 2: Não Encontrados (na lista atual mas não no scraping)
    ws_removidos = wb.create_sheet("Não Encontrados")
    ws_removidos.append(["Ticker"])
    ws_removidos.append(["Total de ativos não encontrados: " + str(len(comparacao['removidos']))])
    ws_removidos.append(["Estes ativos estão na lista atual mas não foram encontrados no scraping"])
    ws_removidos.append([])
    ws_removidos.append(["Ticker"])
    
    # Formatar cabeçalho
    for cell in ws_removidos[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Adicionar tickers não encontrados
    for ticker in sorted(comparacao['removidos']):
        ws_removidos.append([f"{ticker}.sa"])
    
    # Ajustar largura da coluna
    ws_removidos.column_dimensions['A'].width = 15
    
    # Aba 3: Novos (encontrados no scraping mas não na lista atual)
    ws_novos = wb.create_sheet("Novos")
    ws_novos.append(["Ticker"])
    ws_novos.append(["Total de novos ativos encontrados: " + str(len(comparacao['novos']))])
    ws_novos.append(["Estes ativos foram encontrados no scraping mas não estão na lista atual"])
    ws_novos.append([])
    ws_novos.append(["Ticker"])
    
    # Formatar cabeçalho
    for cell in ws_novos[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Adicionar novos tickers
    for ticker in sorted(comparacao['novos']):
        ws_novos.append([f"{ticker}.sa"])
    
    # Ajustar largura da coluna
    ws_novos.column_dimensions['A'].width = 15
    
    # Aba 4: Resumo
    ws_resumo = wb.create_sheet("Resumo", 0)  # Primeira aba
    ws_resumo.append(["COMPARAÇÃO DE AÇÕES - DADOS DE MERCADO"])
    ws_resumo.append([])
    ws_resumo.append(["Data/Hora", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    ws_resumo.append([])
    ws_resumo.append(["Total na lista atual", comparacao['total_atual']])
    ws_resumo.append(["Total encontrado no scraping", comparacao['total_novos']])
    ws_resumo.append(["Ativos em comum", len(comparacao['iguais'])])
    ws_resumo.append(["Novos ativos encontrados", comparacao['total_novos_encontrados']])
    ws_resumo.append(["Ativos não encontrados", comparacao['total_removidos']])
    ws_resumo.append([])
    ws_resumo.append(["Verifique as outras abas para detalhes:"])
    ws_resumo.append(["  - 'Ativos em Comum': Tickers presentes em ambas as listas"])
    ws_resumo.append(["  - 'Não Encontrados': Tickers da lista atual que não foram encontrados"])
    ws_resumo.append(["  - 'Novos': Tickers encontrados no scraping mas não na lista atual"])
    
    # Formatar título
    ws_resumo['A1'].font = Font(bold=True, size=14)
    ws_resumo['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws_resumo['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    
    # Formatar labels
    for row in range(4, 10):
        ws_resumo[f'A{row}'].font = Font(bold=True)
    
    # Ajustar largura das colunas
    ws_resumo.column_dimensions['A'].width = 35
    ws_resumo.column_dimensions['B'].width = 20
    
    # Salvar arquivo
    try:
        wb.save(arquivo_saida)
        return arquivo_saida
    except Exception as e:
        print(f"[ERRO] Erro ao salvar arquivo Excel: {e}")
        return None


def comparar_com_lista_atual(tickers_novos: List[str], arquivo_atual: str = None) -> Dict[str, any]:
    """Compara os tickers extraídos com a lista atual do arquivo assets_lists.py"""
    try:
        # Determinar caminho do arquivo assets_lists.py
        if arquivo_atual is None:
            # Tentar encontrar o arquivo no diretório atual ou no diretório do script
            script_dir = os.path.dirname(os.path.abspath(__file__))
            possiveis_caminhos = [
                os.path.join(script_dir, 'assets_lists.py'),
                os.path.join(os.getcwd(), 'assets_lists.py'),
                'assets_lists.py'
            ]
            
            arquivo_atual = None
            for caminho in possiveis_caminhos:
                if os.path.exists(caminho):
                    arquivo_atual = caminho
                    break
            
            if arquivo_atual is None:
                return {
                    'erro': f'Arquivo assets_lists.py não encontrado. Procurou em: {possiveis_caminhos}'
                }
        
        # Ler arquivo atual
        with open(arquivo_atual, 'r', encoding='utf-8') as f:
            conteudo = f.read()
        
        # Extrair lista atual de ações
        match = re.search(r'LISTA_ACOES\s*=\s*\[(.*?)\]', conteudo, re.DOTALL)
        lista_atual = []
        if match:
            lista_texto = match.group(1)
            # Extrair todos os tickers entre aspas
            tickers_atual = re.findall(r'"([^"]+)"', lista_texto)
            lista_atual = [t.upper().replace('.SA', '').replace('.sa', '') for t in tickers_atual]
        
        # Comparar
        set_atual = set(lista_atual)
        set_novos = set([t.upper().replace('.SA', '').replace('.sa', '') for t in tickers_novos])
        
        novos = sorted(list(set_novos - set_atual))
        removidos = sorted(list(set_atual - set_novos))
        iguais = sorted(list(set_atual & set_novos))
        
        return {
            'total_atual': len(set_atual),
            'total_novos': len(set_novos),
            'novos': novos,
            'removidos': removidos,
            'iguais': iguais,
            'total_novos_encontrados': len(novos),
            'total_removidos': len(removidos)
        }
        
    except Exception as e:
        import traceback
        return {
            'erro': str(e),
            'traceback': traceback.format_exc()
        }


def main():
    """Função principal para testar o scraping do Dados de Mercado"""
    print("=" * 80)
    print("SCRAPING - LISTA DE AÇÕES DO DADOS DE MERCADO")
    print("Busca todas as páginas com paginação")
    print("=" * 80)
    print()
    
    # Extrair ações
    print("[INFO] Iniciando extração de AÇÕES de todas as páginas...")
    print()
    resultado = extrair_acoes_dadosdemercado()
    
    if resultado.get('sucesso'):
        print("\n" + "=" * 80)
        print("RESULTADO FINAL - AÇÕES")
        print("=" * 80)
        print(f"✓ Total de ações encontradas: {resultado['total']}")
        print(f"✓ Páginas processadas: {resultado.get('paginas_processadas', 'N/A')}")
        print(f"\n  Primeiros 20 tickers: {resultado['tickers'][:20]}")
        print(f"  Últimos 20 tickers: {resultado['tickers'][-20:]}")
        
        # Comparar com lista atual
        print("\n" + "-" * 80)
        print("COMPARAÇÃO COM LISTA ATUAL")
        print("-" * 80)
        comparacao = comparar_com_lista_atual(resultado['tickers'])
        if 'erro' not in comparacao:
            print(f"  Total na lista atual: {comparacao['total_atual']}")
            print(f"  Total encontrado no scraping: {comparacao['total_novos']}")
            print(f"  Tickers em comum: {len(comparacao['iguais'])}")
            print(f"\n  ✓ Novos tickers encontrados: {comparacao['total_novos_encontrados']}")
            if comparacao['novos']:
                print(f"  Novos tickers (primeiros 30): {comparacao['novos'][:30]}")
                if len(comparacao['novos']) > 30:
                    print(f"  ... e mais {len(comparacao['novos']) - 30} novos tickers")
            print(f"\n  ⚠ Tickers na lista atual mas não encontrados: {comparacao['total_removidos']}")
            if comparacao['removidos']:
                print(f"  Tickers não encontrados (primeiros 30): {comparacao['removidos'][:30]}")
                if len(comparacao['removidos']) > 30:
                    print(f"  ... e mais {len(comparacao['removidos']) - 30} tickers")
            
            # Gerar arquivo Excel
            print("\n" + "-" * 80)
            print("GERANDO ARQUIVO EXCEL...")
            print("-" * 80)
            arquivo_excel = gerar_excel_comparacao(comparacao, resultado['tickers'])
            if arquivo_excel:
                print(f"  ✓ Arquivo Excel gerado com sucesso: {arquivo_excel}")
                print(f"  ✓ Localização: {os.path.abspath(arquivo_excel)}")
            else:
                print("  ✗ Erro ao gerar arquivo Excel")
        else:
            print(f"  ⚠ Erro na comparação: {comparacao.get('erro')}")
    else:
        print("\n" + "=" * 80)
        print("ERRO")
        print("=" * 80)
        print(f"✗ Erro ao extrair ações: {resultado.get('erro')}")
        if 'traceback' in resultado:
            print("\nTraceback:")
            print(resultado['traceback'])
    
    print("\n" + "=" * 80)
    print("Teste concluído!")
    print("=" * 80)
    
    return resultado


if __name__ == '__main__':
    main()

